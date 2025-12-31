from django.shortcuts import render,redirect,get_object_or_404
import cv2
from PIL import Image
from ultralytics import YOLO  # pyright: ignore[reportMissingImports]
from Accounts.models import CustomUser
import os
import math
import numpy as np
from django.contrib import messages
import matplotlib.pyplot as plt
from math import sqrt
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from pulley_app.forms import ImageUploadForm, Upload_htl_temp
import os
from .models import PulleyDetection
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.hashers import check_password
import time
import threading
from django.http import StreamingHttpResponse, JsonResponse
from django.utils import timezone
from django.conf import settings
from pathlib import Path
import json
from .models import DetectionRecord
from django.db.models import Q
from django.contrib.auth.decorators import login_required
# Create your views here.
# yolo_camera.py

def get_current_temperature_view(request):
    return render(request, 'pulley_app/weather.html')
def location_view(request):
    return render(request, 'pulley_app/location.html')

def get_current_temperature(city):
    url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&appid=2955cdaefeaf8180dffabb6387c0f35a&units=metric"
    response = requests.get(url) # type: ignore
    data = response.json()
    return data['main']['temp']

def main_view(request):
    return render(request, 'pulley_app/index.html')


@login_required(login_url='login')
def railway_view(request):
    return render(request, 'pulley_app/buttons2.html')

# @login_required(login_url='login')
def all_data_view(request):
    if request.user.is_superuser:
        query = request.GET.get('search')
        if query:
            detections = PulleyDetection.objects.filter(
                Q(id__icontains=query) |
                Q(user__icontains=query) |      
                Q(temperature_c__icontains=query) |
                Q(htl_value__icontains=query) |
                Q(total_distance__icontains=query) |
                Q(loss_mm__icontains=query) |
                Q(created_at__icontains=query) |
                Q(email__icontains=query)
            ).order_by('-created_at')
        else:
            detections = PulleyDetection.objects.all().order_by('-created_at')
    else:
        return redirect('login')
    return render(request, 'pulley_app/list_olddata.html', {"detections": detections, "search": query})


def result_data_view(request):
    # if not request.user.is_authenticated:
    #     return redirect('login')
    
    detections = PulleyDetection.objects.filter(user=request.user).order_by('-id')
    print(detections)
    return render(request, 'pulley_app/list_olddata.html', {'detections': detections})


def delete_detections(request, id):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        detection = PulleyDetection.objects.get(id=id)
        detection.delete()
        messages.success(request, "Detection record deleted successfully.")
    except PulleyDetection.DoesNotExist:
        messages.error(request, "Detection record not found.")

    return redirect('old_data')




def all_data_view_for_camera(request):
    detections = []
    if request.user.is_superuser:
        detections = DetectionRecord.objects.all()
    else:
        return redirect('login')
    return render(request, 'pulley_app/data_camera.html', {'detections': detections})

def detete_detections_camera(request, id):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        detection = DetectionRecord.objects.get(id=id)
        detection.delete()
        messages.success(request, "Detection record deleted successfully.")
    except DetectionRecord.DoesNotExist:
        messages.error(request, "Detection record not found.")

    return redirect('all_data_camera')

def result_data_view_for_camera(request):
    detections = DetectionRecord.objects.filter(user=request.user)
    return render(request, 'pulley_app/data_camera.html', {'detections': detections})

def demo_video_view(request):
    return render(request, 'pulley_app/demovideo.html')

def employees_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    else:
        query = request.GET.get('search')
        # if request.user.is_authenticated:
        if query:
            employees = CustomUser.objects.filter( # type: ignore
                Q(id__icontains=query) |
                Q(username__icontains=query) |
                Q(employee_id__icontains=query) |
                Q(email__icontains=query)
            ).order_by('-date_joined').count()

        else:
            employees = CustomUser.objects.all().order_by('-date_joined')

    return render(request, 'pulley_app/userslits2.html', {"employees": employees, "search": query})


@login_required
def employee_deleteview(request, id):
    employee = CustomUser.objects.get(id=id)
    employee.delete()
    messages.success(request, "Employee record deleted successfully.")
    return redirect('employees')

# def employees_view(request):
#     """Employee management page - requires login and staff access"""
#     if not request.user.is_authenticated:
#         return redirect('login')
    
#     # Get all users/employees
#     employees = CustomUser.objects.all().order_by('-date_joined')
    
#     # Count active users
#     active_count = CustomUser.objects.filter(is_active=True).count()
    
#     context = {
#         'employees': employees,
#         'active_count': active_count,
#     }
    
#     return render(request, 'pulley_app/userslits.html', context)


def detect_pulleys(request):
    # Initialize outputs to avoid UnboundLocalError in non-POST or early-exit paths
    out_path = None
    info_lines = []
    result_url = None
    # result_image_url = None
    # distances = []
    if request.method == "POST":
        form = ImageUploadForm(request.POST, request.FILES)
        if form.is_valid():
            image_file = form.cleaned_data["image"]
            print("image_file",image_file)
            # MODEL_PATH = r"D:\PulleyDetector\ai\runs\detect\yolov11m-custom\weights\best.pt"
            MODEL_PATH = r"D:\PulleyDetector\ai\New folder\runs\detect\yolov11m-custom\weights\best.pt"
            
            uploads_storage = FileSystemStorage(
                location=os.path.join(settings.MEDIA_ROOT, 'uploads'),
                base_url=settings.MEDIA_URL + 'uploads/'
            )
            results_storage = FileSystemStorage(
                location=os.path.join(settings.MEDIA_ROOT, 'results'),
                base_url=settings.MEDIA_URL + 'results/'
            )
            filename = uploads_storage.save(image_file.name, image_file)
            IMAGE_PATH = uploads_storage.path(filename)
            # IMAGE_PATH = "D://yolo_model_trian//1.jpg"  # change to your image path
            MM_PER_PIXEL = 1.0  # set your real-world scale (millimeters per pixel)
            # Thermal compensation defaults (total first->third distance)
            BASE_TEMPERATURE_C = 35.0
            STANDARD_TOTAL_DISTANCE_MM = 1300.0  # reference "standard" used for loss computation (1300 - total)
            TEMPERATURE_SENSITIVITY_MM_PER_C = 0.5  # legacy fallback if HTL missing
            HTL_COEFFICIENT = 0.017  # multiplier for HTL-based adjustment
            HTL_MIN = 200.0
            HTL_MAX = 800.0
            # Take temperature from user input; fallback to default if missing
            temp_raw = form.cleaned_data.get("temperature")
            CURRENT_TEMPERATURE_C = float(temp_raw) if temp_raw is not None else 35.0  # °C
            htl_raw = form.cleaned_data.get("htl")
            htl_value = float(htl_raw) if htl_raw is not None else None
            pole_name = form.cleaned_data.get("pole_name", "").strip() or None
            if htl_value is None:
                raise ValueError("HTL value is required.")
            htl_value = float(htl_value)
            if not (HTL_MIN <= htl_value <= HTL_MAX):
                raise ValueError(f"HTL value must be between {HTL_MIN} and {HTL_MAX}.")

            # 2D Chart lookup: Expected total distance (X in mm) based on Temperature and HTL (L/2)
            # Chart structure: CHART_LOOKUP_LOCAL[temp][htl] = X_value_in_mm
            # Based on the adjustment chart for 3:1 ratio pulleys - EXACT VALUES FROM CHART
            CHART_LOOKUP_LOCAL = {
                # Temperature -> HTL -> X value (in mm)
                10: {200: 1385, 225: 1396, 250: 1406, 275: 1417, 300: 1428, 325: 1438, 350: 1449, 375: 1459, 400: 1470, 425: 1481, 450: 1491, 475: 1502, 500: 1513, 525: 1523, 550: 1534, 575: 1544, 600: 1555, 625: 1566, 650: 1576, 675: 1587, 700: 1598, 725: 1608, 750: 1619, 775: 1629, 800: 1640},
                15: {200: 1368, 225: 1377, 250: 1385, 275: 1394, 300: 1402, 325: 1411, 350: 1419, 375: 1428, 400: 1436, 425: 1445, 450: 1453, 475: 1462, 500: 1470, 525: 1479, 550: 1487, 575: 1496, 600: 1504, 625: 1513, 650: 1521, 675: 1530, 700: 1538, 725: 1547, 750: 1555, 775: 1564, 800: 1572},
                20: {200: 1351, 225: 1357, 250: 1364, 275: 1370, 300: 1377, 325: 1383, 350: 1389, 375: 1396, 400: 1402, 425: 1408, 450: 1415, 475: 1421, 500: 1428, 525: 1434, 550: 1440, 575: 1447, 600: 1453, 625: 1459, 650: 1466, 675: 1472, 700: 1479, 725: 1485, 750: 1491, 775: 1498, 800: 1504},
                21: {200: 1348, 225: 1354, 250: 1360, 275: 1365, 300: 1371, 325: 1377, 350: 1383, 375: 1389, 400: 1395, 425: 1401, 450: 1407, 475: 1413, 500: 1419, 525: 1425, 550: 1431, 575: 1437, 600: 1443, 625: 1449, 650: 1455, 675: 1461, 700: 1467, 725: 1473, 750: 1479, 775: 1484, 800: 1490},
                22: {200: 1344, 225: 1350, 250: 1355, 275: 1361, 300: 1366, 325: 1372, 350: 1377, 375: 1383, 400: 1388, 425: 1394, 450: 1399, 475: 1405, 500: 1411, 525: 1416, 550: 1422, 575: 1427, 600: 1433, 625: 1438, 650: 1444, 675: 1449, 700: 1455, 725: 1460, 750: 1466, 775: 1471, 800: 1477},
                23: {200: 1341, 225: 1346, 250: 1351, 275: 1356, 300: 1361, 325: 1366, 350: 1371, 375: 1377, 400: 1382, 425: 1387, 450: 1392, 475: 1397, 500: 1402, 525: 1407, 550: 1412, 575: 1417, 600: 1422, 625: 1428, 650: 1433, 675: 1438, 700: 1443, 725: 1448, 750: 1453, 775: 1458, 800: 1463},
                24: {200: 1337, 225: 1342, 250: 1347, 275: 1351, 300: 1356, 325: 1361, 350: 1365, 375: 1370, 400: 1375, 425: 1379, 450: 1384, 475: 1389, 500: 1394, 525: 1398, 550: 1403, 575: 1408, 600: 1412, 625: 1417, 650: 1422, 675: 1426, 700: 1431, 725: 1436, 750: 1440, 775: 1445, 800: 1450},
                25: {200: 1334, 225: 1338, 250: 1343, 275: 1347, 300: 1351, 325: 1355, 350: 1360, 375: 1364, 400: 1368, 425: 1372, 450: 1377, 475: 1381, 500: 1385, 525: 1389, 550: 1394, 575: 1398, 600: 1402, 625: 1406, 650: 1411, 675: 1415, 700: 1419, 725: 1423, 750: 1428, 775: 1432, 800: 1436},
                26: {200: 1331, 225: 1334, 250: 1338, 275: 1342, 300: 1346, 325: 1350, 350: 1354, 375: 1357, 400: 1361, 425: 1365, 450: 1369, 475: 1373, 500: 1377, 525: 1380, 550: 1384, 575: 1388, 600: 1392, 625: 1396, 650: 1399, 675: 1403, 700: 1407, 725: 1411, 750: 1415, 775: 1419, 800: 1422},
                27: {200: 1327, 225: 1331, 250: 1334, 275: 1337, 300: 1341, 325: 1344, 350: 1348, 375: 1351, 400: 1354, 425: 1358, 450: 1361, 475: 1365, 500: 1368, 525: 1371, 550: 1375, 575: 1378, 600: 1382, 625: 1385, 650: 1388, 675: 1392, 700: 1395, 725: 1399, 750: 1402, 775: 1405, 800: 1409},
                28: {200: 1324, 225: 1327, 250: 1330, 275: 1333, 300: 1336, 325: 1339, 350: 1342, 375: 1345, 400: 1348, 425: 1351, 450: 1354, 475: 1357, 500: 1360, 525: 1362, 550: 1365, 575: 1368, 600: 1371, 625: 1374, 650: 1377, 675: 1380, 700: 1383, 725: 1386, 750: 1389, 775: 1392, 800: 1395},
                29: {200: 1320, 225: 1323, 250: 1326, 275: 1328, 300: 1331, 325: 1333, 350: 1336, 375: 1338, 400: 1341, 425: 1343, 450: 1346, 475: 1348, 500: 1351, 525: 1354, 550: 1356, 575: 1359, 600: 1361, 625: 1364, 650: 1366, 675: 1369, 700: 1371, 725: 1374, 750: 1377, 775: 1379, 800: 1382},
                30: {200: 1317, 225: 1319, 250: 1321, 275: 1323, 300: 1326, 325: 1328, 350: 1330, 375: 1332, 400: 1334, 425: 1336, 450: 1338, 475: 1340, 500: 1343, 525: 1345, 550: 1347, 575: 1349, 600: 1351, 625: 1353, 650: 1355, 675: 1357, 700: 1360, 725: 1362, 750: 1364, 775: 1366, 800: 1368},
                31: {200: 1314, 225: 1315, 250: 1317, 275: 1319, 300: 1320, 325: 1322, 350: 1324, 375: 1326, 400: 1327, 425: 1329, 450: 1331, 475: 1332, 500: 1334, 525: 1336, 550: 1337, 575: 1339, 600: 1341, 625: 1343, 650: 1344, 675: 1346, 700: 1348, 725: 1349, 750: 1351, 775: 1353, 800: 1354},
                32: {200: 1310, 225: 1311, 250: 1313, 275: 1314, 300: 1315, 325: 1317, 350: 1318, 375: 1319, 400: 1320, 425: 1322, 450: 1323, 475: 1324, 500: 1326, 525: 1327, 550: 1328, 575: 1329, 600: 1331, 625: 1332, 650: 1333, 675: 1334, 700: 1336, 725: 1337, 750: 1338, 775: 1340, 800: 1341},
                33: {200: 1307, 225: 1308, 250: 1309, 275: 1309, 300: 1310, 325: 1311, 350: 1312, 375: 1313, 400: 1314, 425: 1314, 450: 1315, 475: 1316, 500: 1317, 525: 1318, 550: 1319, 575: 1320, 600: 1320, 625: 1321, 650: 1322, 675: 1323, 700: 1324, 725: 1325, 750: 1326, 775: 1326, 800: 1327},
                34: {200: 1303, 225: 1304, 250: 1304, 275: 1305, 300: 1305, 325: 1306, 350: 1306, 375: 1306, 400: 1307, 425: 1307, 450: 1308, 475: 1308, 500: 1309, 525: 1309, 550: 1309, 575: 1310, 600: 1310, 625: 1311, 650: 1311, 675: 1311, 700: 1312, 725: 1312, 750: 1313, 775: 1313, 800: 1314},
                35: {200: 1300, 225: 1300, 250: 1300, 275: 1300, 300: 1300, 325: 1300, 350: 1300, 375: 1300, 400: 1300, 425: 1300, 450: 1300, 475: 1300, 500: 1300, 525: 1300, 550: 1300, 575: 1300, 600: 1300, 625: 1300, 650: 1300, 675: 1300, 700: 1300, 725: 1300, 750: 1300, 775: 1300, 800: 1300},
                36: {200: 1297, 225: 1296, 250: 1296, 275: 1295, 300: 1295, 325: 1294, 350: 1294, 375: 1294, 400: 1293, 425: 1293, 450: 1292, 475: 1292, 500: 1292, 525: 1291, 550: 1291, 575: 1290, 600: 1290, 625: 1289, 650: 1289, 675: 1289, 700: 1288, 725: 1288, 750: 1287, 775: 1287, 800: 1286},
                37: {200: 1293, 225: 1292, 250: 1292, 275: 1291, 300: 1290, 325: 1289, 350: 1288, 375: 1287, 400: 1286, 425: 1286, 450: 1285, 475: 1284, 500: 1283, 525: 1282, 550: 1281, 575: 1280, 600: 1280, 625: 1279, 650: 1278, 675: 1277, 700: 1276, 725: 1275, 750: 1275, 775: 1274, 800: 1273},
                38: {200: 1290, 225: 1289, 250: 1287, 275: 1286, 300: 1285, 325: 1283, 350: 1282, 375: 1281, 400: 1280, 425: 1278, 450: 1277, 475: 1276, 500: 1275, 525: 1273, 550: 1272, 575: 1271, 600: 1269, 625: 1268, 650: 1267, 675: 1266, 700: 1264, 725: 1263, 750: 1262, 775: 1260, 800: 1259},
                39: {200: 1286, 225: 1285, 250: 1283, 275: 1281, 300: 1280, 325: 1278, 350: 1276, 375: 1275, 400: 1273, 425: 1271, 450: 1269, 475: 1268, 500: 1266, 525: 1264, 550: 1263, 575: 1261, 600: 1259, 625: 1258, 650: 1256, 675: 1254, 700: 1252, 725: 1251, 750: 1249, 775: 1247, 800: 1246},
                40: {200: 1283, 225: 1281, 250: 1279, 275: 1277, 300: 1275, 325: 1272, 350: 1270, 375: 1268, 400: 1266, 425: 1264, 450: 1262, 475: 1260, 500: 1258, 525: 1255, 550: 1253, 575: 1251, 600: 1249, 625: 1247, 650: 1245, 675: 1243, 700: 1241, 725: 1238, 750: 1236, 775: 1234, 800: 1232},
                41: {200: 1280, 225: 1277, 250: 1275, 275: 1272, 300: 1269, 325: 1267, 350: 1264, 375: 1262, 400: 1259, 425: 1257, 450: 1254, 475: 1252, 500: 1249, 525: 1246, 550: 1244, 575: 1241, 600: 1239, 625: 1236, 650: 1234, 675: 1231, 700: 1229, 725: 1226, 750: 1224, 775: 1221, 800: 1218},
                42: {200: 1276, 225: 1273, 250: 1270, 275: 1267, 300: 1264, 325: 1261, 350: 1258, 375: 1255, 400: 1252, 425: 1249, 450: 1246, 475: 1243, 500: 1241, 525: 1238, 550: 1235, 575: 1232, 600: 1229, 625: 1226, 650: 1223, 675: 1220, 700: 1217, 725: 1214, 750: 1211, 775: 1208, 800: 1205},
                43: {200: 1273, 225: 1269, 250: 1266, 275: 1263, 300: 1259, 325: 1256, 350: 1252, 375: 1249, 400: 1246, 425: 1242, 450: 1239, 475: 1235, 500: 1232, 525: 1229, 550: 1225, 575: 1222, 600: 1218, 625: 1215, 650: 1212, 675: 1208, 700: 1205, 725: 1201, 750: 1198, 775: 1195, 800: 1191},
                44: {200: 1269, 225: 1266, 250: 1262, 275: 1258, 300: 1254, 325: 1250, 350: 1246, 375: 1243, 400: 1239, 425: 1235, 450: 1231, 475: 1227, 500: 1224, 525: 1220, 550: 1216, 575: 1212, 600: 1208, 625: 1204, 650: 1201, 675: 1197, 700: 1193, 725: 1189, 750: 1185, 775: 1181, 800: 1178},
                45: {200: 1266, 225: 1262, 250: 1258, 275: 1253, 300: 1249, 325: 1245, 350: 1241, 375: 1236, 400: 1232, 425: 1228, 450: 1224, 475: 1219, 500: 1215, 525: 1211, 550: 1207, 575: 1202, 600: 1198, 625: 1194, 650: 1190, 675: 1185, 700: 1181, 725: 1177, 750: 1173, 775: 1168, 800: 1164},
                50: {200: 1249, 225: 1243, 250: 1236, 275: 1230, 300: 1224, 325: 1217, 350: 1211, 375: 1204, 400: 1198, 425: 1192, 450: 1185, 475: 1179, 500: 1173, 525: 1166, 550: 1160, 575: 1153, 600: 1147, 625: 1141, 650: 1134, 675: 1128, 700: 1122, 725: 1115, 750: 1109, 775: 1102, 800: 1096},
            }

            # Ratio assumptions for a 3:1 pulley layout (first->second, second->third)
            DISTANCE_RATIO_12 = 0.75
            DISTANCE_RATIO_23 = 0.25
            # NUMBERING_SIDE = "left"  # choose "left" (default) or "right" to start numbering from that side


            model = YOLO(MODEL_PATH)

            results = model.predict(source=IMAGE_PATH, save=True, verbose=False)
            if not results:
                raise RuntimeError("No results returned by the model.")

            result = results[0]
            if result.boxes is None or len(result.boxes) == 0:
                raise RuntimeError("No detections found in the image.")

            names = result.names if hasattr(result, "names") else model.names
            boxes = result.boxes
            cls_indices = boxes.cls.cpu().numpy().astype(int)
            xyxy = boxes.xyxy.cpu().numpy()

            # Collect pulley centers
            pulley_points = []
            for i, c in enumerate(cls_indices):
                label = names.get(int(c), str(c)) if isinstance(names, dict) else names[int(c)]
                if str(label).lower() == "pulley":
                    x1, y1, x2, y2 = xyxy[i]
                    cx = (x1 + x2) / 2.0
                    cy = (y1 + y2) / 2.0
                    pulley_points.append((cx, cy))

            if len(pulley_points) < 2:
                raise RuntimeError("Found fewer than 2 pulleys. Need at least 2 to compute distance.")

            # Sort left-to-right to define first, second, third
            pulley_points.sort(key=lambda p: p[0])
            p1 = pulley_points[0] if len(pulley_points) >= 1 else None
            p2 = pulley_points[1] if len(pulley_points) >= 2 else None
            p3 = pulley_points[2] if len(pulley_points) >= 3 else None

            # Load image for drawing
            img = cv2.imread(IMAGE_PATH)
            if img is None:
                raise RuntimeError(f"Failed to load image: {IMAGE_PATH}")

            # Draw points
            for (cx, cy) in pulley_points:
                cv2.circle(img, (int(cx), int(cy)), 6, (0, 0, 255), -1)
            # Draw pulley indices (1, 2, 3) next to centers
            for idx, (cx, cy) in enumerate(pulley_points, start=1):
                cv2.putText(
                    img,
                    f"{idx}",
                    (int(cx) + 8, int(cy) - 8),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.9,
                    (0, 0, 255),
                    2,
                    cv2.LINE_AA
                )

            def pixel_distance(p1, p2):
                return math.hypot(p1[0] - p2[0], p1[1] - p2[1])

            def _chart_total_distance(temp_c):
                """
                Return chart-based total distance X for temperature (with linear interpolation).
                If the chart has no data, return None.
                """
                if not TEMP_TO_TOTAL_DISTANCE_MM: # type: ignore
                    return None
                # Exact hit
                if temp_c in TEMP_TO_TOTAL_DISTANCE_MM: # type: ignore
                    return float(TEMP_TO_TOTAL_DISTANCE_MM[temp_c]) # type: ignore
                # Find neighbors for interpolation
                keys = sorted(TEMP_TO_TOTAL_DISTANCE_MM.keys()) # type: ignore
                if temp_c <= keys[0]:
                    return float(TEMP_TO_TOTAL_DISTANCE_MM[keys[0]]) # type: ignore
                if temp_c >= keys[-1]:
                    return float(TEMP_TO_TOTAL_DISTANCE_MM[keys[-1]]) # type: ignore
                lower = None
                upper = None
                for k in keys:
                    if k < temp_c:
                        lower = k
                    elif k > temp_c:
                        upper = k
                        break
                if lower is None or upper is None:
                    return None
                x0 = lower
                y0 = float(TEMP_TO_TOTAL_DISTANCE_MM[lower]) # type: ignore
                x1 = upper
                y1 = float(TEMP_TO_TOTAL_DISTANCE_MM[upper]) # type: ignore
                ratio = (temp_c - x0) / (x1 - x0)
                return y0 + ratio * (y1 - y0)

            def expected_total_distance_for_temperature(temp_c, htl=None):
                """
                Compute expected total pulley distance (1->3) in mm for the given temperature.
                Priority: use HTL-based formula. If unavailable, use chart values, otherwise linear fallback.
                """
                if htl is not None:
                    delta_t = temp_c - BASE_TEMPERATURE_C
                    adjustment = (htl * HTL_COEFFICIENT) * abs(delta_t)
                    if delta_t > 0:
                        return STANDARD_TOTAL_DISTANCE_MM - adjustment
                    if delta_t < 0:
                        return STANDARD_TOTAL_DISTANCE_MM + adjustment
                    return STANDARD_TOTAL_DISTANCE_MM

                chart_value = _chart_total_distance(temp_c)
                if chart_value is not None:
                    return chart_value
                delta_t = temp_c - BASE_TEMPERATURE_C
                return STANDARD_TOTAL_DISTANCE_MM - TEMPERATURE_SENSITIVITY_MM_PER_C * delta_t

            def temperature_from_total_distance(distance_mm):
                """
                Estimate temperature (°C) from a measured total pulley distance 1->3.
                """
                delta_d = STANDARD_TOTAL_DISTANCE_MM - distance_mm
                return BASE_TEMPERATURE_C + (delta_d / TEMPERATURE_SENSITIVITY_MM_PER_C)

            def split_total_distance(distance_mm):
                """
                Split total 1->3 distance into 1->2 and 2->3 spans using configured ratios.
                """
                return (
                    distance_mm * DISTANCE_RATIO_12,
                    distance_mm * DISTANCE_RATIO_23
                )

            dist12_mm = None
            dist23_mm = None

            # Compute and draw 1->2
            if p1 is not None and p2 is not None:
                d_px = pixel_distance(p1, p2)
                #3x left-site
                # dist12_mm = 1.8*d_px * MM_PER_PIXEL +25
                #3x right-site
                dist12_mm = 2.3*d_px * MM_PER_PIXEL
                cv2.line(img, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), (0, 0,255), 2)
                mid12 = (int((p1[0] + p2[0]) / 2), int((p1[1] + p2[1]) / 2))
                # cv2.putText(img, f"{dist12_mm:.2f} mm", mid12, cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv2.LINE_AA)
                cv2.putText(img, f"", mid12, cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv2.LINE_AA)
                
            

            # Compute and draw 2->3
            total_distance_mm = None

            if p2 is not None and p3 is not None:
                d_px = pixel_distance(p2, p3)
                #3x left-site
                # dist23_mm = 2.3*d_px * MM_PER_PIXEL
                #4x
                dist23_mm = 2.4*d_px * MM_PER_PIXEL+15
                cv2.line(img, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), 2)
                mid23 = (int((p2[0] + p3[0]) / 2), int((p2[1] + p3[1]) / 2))
                cv2.putText(img, f"{dist23_mm:.2f} mm", mid23, cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2, cv2.LINE_AA)

            if dist12_mm is not None and dist23_mm is not None:
                # total_distance_mm = dist12_mm + dist23_mm
                total_distance_mm =  dist23_mm
                

            expected_total_db = None
            loss_mm_db = None

            if total_distance_mm is not None and TEMPERATURE_SENSITIVITY_MM_PER_C > 0 and p1 is not None and p3 is not None:
                expected_total = expected_total_distance_for_temperature(CURRENT_TEMPERATURE_C, htl_value)
                expected_dist12, expected_dist23 = split_total_distance(expected_total)
                estimated_temp = temperature_from_total_distance(total_distance_mm)
                loss_mm = expected_total - total_distance_mm
                expected_total_db = expected_total
                loss_mm_db = loss_mm
                info_lines = [
                    f"Temperature: {CURRENT_TEMPERATURE_C:.1f} °C",
                    f"HTL (L/2): {htl_value:.1f}",
                    # f"Pulley 1->Pulley 2: {dist12_mm:.3f} mm",
                    f"Pulley 2->Pulley 3: {dist23_mm:.3f} mm",
                    # f"Total distance (1->3): {total_distance_mm:.3f} mm",
                    f"Expected @ {CURRENT_TEMPERATURE_C:.1f} °C (HTL {htl_value:.1f}): {expected_total:.3f} mm",
                    # f"Expected 1->2: {expected_dist12:.3f} mm | 2->3: {expected_dist23:.3f} mm",
                    # f"Expected  2->3: {expected_dist23:.3f} mm",
                    f"Loss vs expected:{CURRENT_TEMPERATURE_C:.1f} °C: {expected_total:.3f} mm - {dist23_mm:.3f} mm = {loss_mm:.3f} mm",
                    # f"Estimated temperature: {estimated_temp:.2f} C"
                ]
                text_y = int(min(p1[1], p3[1]) - 50)
                for line in info_lines:
                    cv2.putText(
                        img,
                        line,
                        (int((p1[0] + p3[0]) / 2 - 160), max(text_y, 30)),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.6,
                        (255, 0, 0),
                        2,
                        cv2.LINE_AA
                    )
                    text_y -= 25

            # Save annotated image under results folder using uploaded filename as base
            uploaded_root, uploaded_ext = os.path.splitext(filename)
            result_filename = f"{uploaded_root}_output{uploaded_ext}"
            out_path = results_storage.path(result_filename)
            print(f"DEBUG: Saving image to: {out_path}")
            cv2.imwrite(out_path, img)
            # Verify file was saved
            if os.path.exists(out_path):
                print(f"DEBUG: Image saved successfully at: {out_path}")
            else:
                print(f"DEBUG: ERROR - Image file not found at: {out_path}")
            # Public URL for template rendering
            result_url = results_storage.url(result_filename)
            print(f"DEBUG: Result image URL: {result_url}")
            print(f"DEBUG: Full URL would be: {request.build_absolute_uri(result_url)}")
            # cv2.imshow("output",img)
            # cv2.waitKey(0)
            # cv2.destroyAllWindows()
        

            # Print distances
            print(f"HTL (L/2) value: {htl_value:.1f}")
            if dist12_mm is not None:
                print(f"Distance first->second pulley: {dist12_mm:.3f} mm (scale {MM_PER_PIXEL} mm/px)")
            if dist23_mm is not None:
                print(f"Distance second->third pulley: {dist23_mm:.3f} mm (scale {MM_PER_PIXEL} mm/px)")
            if total_distance_mm is not None and TEMPERATURE_SENSITIVITY_MM_PER_C > 0:
                expected_total = expected_total_distance_for_temperature(CURRENT_TEMPERATURE_C, htl_value)
                expected_dist12, expected_dist23 = split_total_distance(expected_total)
                estimated_temp = temperature_from_total_distance(total_distance_mm)
                loss_mm = expected_total - total_distance_mm
                print(f"Expected total distance (1->3) at {CURRENT_TEMPERATURE_C:.1f} °C with HTL {htl_value:.1f}: {expected_total:.3f} mm")
                print(f"Expected split: 1->2 = {expected_dist12:.3f} mm, 2->3 = {expected_dist23:.3f} mm")
                print(f"Measured total distance (1->3): {total_distance_mm:.3f} mm")
                print(f"Loss vs expected:{CURRENT_TEMPERATURE_C:.1f} °C: {expected_total:.3f} mm - {total_distance_mm:.3f} mm = {loss_mm:.3f} mm")
                # print(f"Estimated temperature from measured total: {estimated_temp:.3f} °C")
            if len(pulley_points) < 3:
                print("Only two pulleys detected; third-to-second distance not computed.")
            print(f"Annotated image saved to: {out_path}")

            # Save to model
            try:
                distance_summary = ""
                if dist12_mm is not None:
                    distance_summary += f"Pulley 1->2: {dist12_mm:.3f} mm | "
                if dist23_mm is not None:
                    distance_summary += f"Pulley 2->3: {dist23_mm:.3f} mm | "
                if total_distance_mm is not None:
                    distance_summary += f"Total: {total_distance_mm:.3f} mm | "
                if expected_total_db is not None:
                    distance_summary += (
                        f"Standard (HTL {htl_value:.0f}) @ {CURRENT_TEMPERATURE_C:.1f}°C: {expected_total_db:.3f} mm | "
                        f"Diff: {loss_mm_db:.3f} mm"
                    )
                distance_summary = distance_summary.rstrip(" | ")

                PulleyDetection.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    pole_name=pole_name,
                    uploaded_image=f"uploads/{filename}",
                    result_image=f"results/{result_filename}",
                    temperature_c=CURRENT_TEMPERATURE_C,
                    htl_value=htl_value,
                    dist_p1_p2=dist12_mm,
                    dist_p2_p3=dist23_mm,
                    total_distance=total_distance_mm,
                    expected_total=expected_total_db,
                    loss_mm=loss_mm_db,
                    distances=distance_summary
                )
                
                
            
            except Exception as e:
                print(f"Warning: failed to save PulleyDetection record: {e}")
    else:
        form = ImageUploadForm()

    return render(request, "pulley_app/upload.html", {
        "form": form,
        "result_image_url": result_url,
        "distances": info_lines,
    })

# Global storage for detection results (thread-safe with locks)
detection_data = {
    'dist12': None,
    'dist23': None,
    'total': None,
    'points': [],
    'confidences': [],
    'expected_total': None,
    'expected_dist12': None,
    'expected_dist23': None,
    'loss_mm': None,
    'pulley_count': 0,
    'frame_available': False,
    'last_update': None,
    'latest_frame': None
}
detection_lock = threading.Lock()
camera_cap = None
camera_running = False

# def bookings_view(request):
#         if not request.user.is_authenticated:
#             return redirect('login')
#         else:
#             return render(request,'pulley_app/bookings.html')
#         # response = render(request, 'pulley_app/bookings.html')
#         # response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
#         # response['Pragma'] = 'no-cache'
#         # response['Expires'] = '0'
#         # return response

def services_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    response = render(request, 'pulley_app/services.html')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

def support_view(request):
    """Support and Settings page - requires login"""
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'pulley_app/support.html')

def calculator_view(request):
    """Calculator page - requires login"""
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'pulley_app/calculator.html')
def calculator_buttons_view(request):
    """Calculator Buttons page - requires login"""
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'pulley_app/calcbuttons.html')
def pulley_calculator_views(request):
    return render(request, 'pulley_app/pulley_calculator.html')


def chooes_your_database_view(request):
    return render(request, 'pulley_app/buttonimg.html')
def chooes_your_database_view2(request):
    return render(request, 'pulley_app/buttonimg2.html')

#88888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888
#second sections------------------------------------------------------------------------------
# optonal second part of project-----------------------------------------------------------------------------------------
# yolo_camera.py
CAPTURE_SUBDIR = Path("best_captures")
MEDIA_ROOT = Path(getattr(settings, "MEDIA_ROOT", Path(settings.BASE_DIR) / "media"))
BEST_CAPTURE_DIR = MEDIA_ROOT / CAPTURE_SUBDIR
REFERENCE_LABELS = ("poll", "pole", "counter weight")
MODEL_PATH = Path(r"D:\PulleyDetector\ai\runs\detect\yolov11m-custom\weights\best.pt")
_yolo_model = None
_model_lock = threading.Lock()


def _default_detection_state():
    return {
        'dist12': None,
        'dist23': None,
        'total': None,
        'points': [],
        'confidences': [],
        'segments': [],
        'expected_total': None,
        'expected_dist12': None,
        'expected_dist23': None,
        'loss_mm': None,
        'pulley_count': 0,
        'temperature_c': None,
        'htl_value': None,
        'frame_available': False,
        'last_update': None,
        'latest_frame': None,
        'capture_complete': False,
        'capture_image_path': None,
        'capture_requested': False,
    }


# Global storage for detection results (thread-safe with locks)
detection_data = _default_detection_state()
detection_lock = threading.Lock()
camera_cap = None
camera_running = False


def reset_detection_state():
    """Return detection storage to its initial blank state."""
    with detection_lock:
        detection_data.clear()
        detection_data.update(_default_detection_state())


def stop_camera_feed():
    """Stop and release the camera safely."""
    global camera_cap, camera_running
    with detection_lock:
        camera_running = False
        if camera_cap is not None:
            try:
                camera_cap.release()
            except Exception:
                pass
            camera_cap = None


def _get_yolo_model():
    global _yolo_model
    with _model_lock:
        if _yolo_model is None:
            _yolo_model = YOLO(str(MODEL_PATH))
    return _yolo_model


def save_detection_frame(frame):
    """Persist an annotated frame to disk and return the media-relative path."""
    if frame is None:
        return None
    try:
        BEST_CAPTURE_DIR.mkdir(parents=True, exist_ok=True)
        filename = f"detection_{int(time.time())}.jpg"
        file_path = BEST_CAPTURE_DIR / filename
        cv2.imwrite(str(file_path), frame)
        return (CAPTURE_SUBDIR / filename).as_posix()
    except Exception as exc:
        print(f"Error saving detection frame: {exc}")
        return None

def index(request):
    return render(request, 'index.html')


def forms_view(request):
    temp = request.session.get('live_temperature_c')
    htl = request.session.get('live_htl_value')
    pole_name = request.session.get('live_pole_name')
    if request.method == 'POST':
        form = Upload_htl_temp(request.POST)
        if form.is_valid():
            temp = float(form.cleaned_data.get('temperature'))
            htl = float(form.cleaned_data.get('htl'))
            pole_name = form.cleaned_data.get('pole_name', '').strip() or None
            request.session['live_temperature_c'] = temp
            request.session['live_htl_value'] = htl
            request.session['live_pole_name'] = pole_name
            messages.success(request, "Saved temperature and HTL. Starting live detection...")
            return redirect('yolo_camera')
    else:
        initial = {}
        if temp is not None:
            initial['temperature'] = str(temp)
        if htl is not None:
            initial['htl'] = str(htl)
        if pole_name is not None:
            initial['pole_name'] = pole_name
        form = Upload_htl_temp(initial=initial or None)

    return render(request, 'pulley_app/forms.html', {'form': form,'temp':temp,'htl':htl})
def yolo_camera(request):
    if request.GET.get('restart') == '1':
        stop_camera_feed()
        existing_thread = getattr(yolo_camera, "_detection_thread", None)
        if existing_thread and existing_thread.is_alive():
            try:
                existing_thread.join(timeout=0.5)
            except Exception:
                pass
        reset_detection_state()
        yolo_camera._detection_thread = None

    # Require temperature and HTL before starting live detection
    temp_param = request.GET.get('temperature') or request.session.get('live_temperature_c')
    htl_param = request.GET.get('htl') or request.session.get('live_htl_value')
    pole_name_param = request.GET.get('pole_name') or request.session.get('live_pole_name')
    if temp_param is None or htl_param is None:
        messages.error(request, "Please select temperature and HTL before starting live detection.")
        return redirect('forms')
    try:
        CURRENT_TEMPERATURE_C = float(temp_param)
    except (TypeError, ValueError):
        messages.error(request, "Invalid temperature value. Please try again.")
        return redirect('forms')
    try:
        htl_value = float(htl_param)
    except (TypeError, ValueError):
        messages.error(request, "Invalid HTL value. Please try again.")
        return redirect('forms')

    request.session['live_temperature_c'] = CURRENT_TEMPERATURE_C
    request.session['live_htl_value'] = htl_value
    # Persist pole name consistently in session for live detections
    if pole_name_param:
        request.session['live_pole_name'] = pole_name_param.strip() or None
    else:
        request.session['live_pole_name'] = None
    pole_name = request.session.get('live_pole_name')
    detection_user = request.user if request.user.is_authenticated else None

    MM_PER_PIXEL = 1.0  # update with your calibration value
    CONF_THRES = 0.25
    IOU_THRES = 0.45
    IMAGE_SZ = 640

    # Thermal / expected distance configuration
    BASE_TEMPERATURE_C = 35.0
    STANDARD_TOTAL_DISTANCE_MM = 1300.0
    TEMPERATURE_SENSITIVITY_MM_PER_C = 0.5
    # CURRENT_TEMPERATURE_C set from user/session above

    # Ratio assumptions for a 3:1 pulley layout (first->second, second->third)
    DISTANCE_RATIO_12 = 0.75
    DISTANCE_RATIO_23 = 0.25

    # 2D Chart lookup: Expected total distance (X in mm) based on Temperature and HTL (L/2)
    # Chart structure: CHART_LOOKUP[temp][htl] = X_value_in_mm
    # Based on the adjustment chart for 3:1 ratio pulleys - EXACT VALUES FROM CHART
    CHART_LOOKUP = {
        # Temperature -> HTL -> X value (in mm)
        10: {200: 1385, 225: 1396, 250: 1406, 275: 1417, 300: 1428, 325: 1438, 350: 1449, 375: 1459, 400: 1470, 425: 1481, 450: 1491, 475: 1502, 500: 1513, 525: 1523, 550: 1534, 575: 1544, 600: 1555, 625: 1566, 650: 1576, 675: 1587, 700: 1598, 725: 1608, 750: 1619, 775: 1629, 800: 1640},
        15: {200: 1368, 225: 1377, 250: 1385, 275: 1394, 300: 1402, 325: 1411, 350: 1419, 375: 1428, 400: 1436, 425: 1445, 450: 1453, 475: 1462, 500: 1470, 525: 1479, 550: 1487, 575: 1496, 600: 1504, 625: 1513, 650: 1521, 675: 1530, 700: 1538, 725: 1547, 750: 1555, 775: 1564, 800: 1572},
        20: {200: 1351, 225: 1357, 250: 1364, 275: 1370, 300: 1377, 325: 1383, 350: 1389, 375: 1396, 400: 1402, 425: 1408, 450: 1415, 475: 1421, 500: 1428, 525: 1434, 550: 1440, 575: 1447, 600: 1453, 625: 1459, 650: 1466, 675: 1472, 700: 1479, 725: 1485, 750: 1491, 775: 1498, 800: 1504},
        21: {200: 1348, 225: 1354, 250: 1360, 275: 1365, 300: 1371, 325: 1377, 350: 1383, 375: 1389, 400: 1395, 425: 1401, 450: 1407, 475: 1413, 500: 1419, 525: 1425, 550: 1431, 575: 1437, 600: 1443, 625: 1449, 650: 1455, 675: 1461, 700: 1467, 725: 1473, 750: 1479, 775: 1484, 800: 1490},
        22: {200: 1344, 225: 1350, 250: 1355, 275: 1361, 300: 1366, 325: 1372, 350: 1377, 375: 1383, 400: 1388, 425: 1394, 450: 1399, 475: 1405, 500: 1411, 525: 1416, 550: 1422, 575: 1427, 600: 1433, 625: 1438, 650: 1444, 675: 1449, 700: 1455, 725: 1460, 750: 1466, 775: 1471, 800: 1477},
        23: {200: 1341, 225: 1346, 250: 1351, 275: 1356, 300: 1361, 325: 1366, 350: 1371, 375: 1377, 400: 1382, 425: 1387, 450: 1392, 475: 1397, 500: 1402, 525: 1407, 550: 1412, 575: 1417, 600: 1422, 625: 1428, 650: 1433, 675: 1438, 700: 1443, 725: 1448, 750: 1453, 775: 1458, 800: 1463},
        24: {200: 1337, 225: 1342, 250: 1347, 275: 1351, 300: 1356, 325: 1361, 350: 1365, 375: 1370, 400: 1375, 425: 1379, 450: 1384, 475: 1389, 500: 1394, 525: 1398, 550: 1403, 575: 1408, 600: 1412, 625: 1417, 650: 1422, 675: 1426, 700: 1431, 725: 1436, 750: 1440, 775: 1445, 800: 1450},
        25: {200: 1334, 225: 1338, 250: 1343, 275: 1347, 300: 1351, 325: 1355, 350: 1360, 375: 1364, 400: 1368, 425: 1372, 450: 1377, 475: 1381, 500: 1385, 525: 1389, 550: 1394, 575: 1398, 600: 1402, 625: 1406, 650: 1411, 675: 1415, 700: 1419, 725: 1423, 750: 1428, 775: 1432, 800: 1436},
        26: {200: 1331, 225: 1334, 250: 1338, 275: 1342, 300: 1346, 325: 1350, 350: 1354, 375: 1357, 400: 1361, 425: 1365, 450: 1369, 475: 1373, 500: 1377, 525: 1380, 550: 1384, 575: 1388, 600: 1392, 625: 1396, 650: 1399, 675: 1403, 700: 1407, 725: 1411, 750: 1415, 775: 1419, 800: 1422},
        27: {200: 1327, 225: 1331, 250: 1334, 275: 1337, 300: 1341, 325: 1344, 350: 1348, 375: 1351, 400: 1354, 425: 1358, 450: 1361, 475: 1365, 500: 1368, 525: 1371, 550: 1375, 575: 1378, 600: 1382, 625: 1385, 650: 1388, 675: 1392, 700: 1395, 725: 1399, 750: 1402, 775: 1405, 800: 1409},
        28: {200: 1324, 225: 1327, 250: 1330, 275: 1333, 300: 1336, 325: 1339, 350: 1342, 375: 1345, 400: 1348, 425: 1351, 450: 1354, 475: 1357, 500: 1360, 525: 1362, 550: 1365, 575: 1368, 600: 1371, 625: 1374, 650: 1377, 675: 1380, 700: 1383, 725: 1386, 750: 1389, 775: 1392, 800: 1395},
        29: {200: 1320, 225: 1323, 250: 1326, 275: 1328, 300: 1331, 325: 1333, 350: 1336, 375: 1338, 400: 1341, 425: 1343, 450: 1346, 475: 1348, 500: 1351, 525: 1354, 550: 1356, 575: 1359, 600: 1361, 625: 1364, 650: 1366, 675: 1369, 700: 1371, 725: 1374, 750: 1377, 775: 1379, 800: 1382},
        30: {200: 1317, 225: 1319, 250: 1321, 275: 1323, 300: 1326, 325: 1328, 350: 1330, 375: 1332, 400: 1334, 425: 1336, 450: 1338, 475: 1340, 500: 1343, 525: 1345, 550: 1347, 575: 1349, 600: 1351, 625: 1353, 650: 1355, 675: 1357, 700: 1360, 725: 1362, 750: 1364, 775: 1366, 800: 1368},
        31: {200: 1314, 225: 1315, 250: 1317, 275: 1319, 300: 1320, 325: 1322, 350: 1324, 375: 1326, 400: 1327, 425: 1329, 450: 1331, 475: 1332, 500: 1334, 525: 1336, 550: 1337, 575: 1339, 600: 1341, 625: 1343, 650: 1344, 675: 1346, 700: 1348, 725: 1349, 750: 1351, 775: 1353, 800: 1354},
        32: {200: 1310, 225: 1311, 250: 1313, 275: 1314, 300: 1315, 325: 1317, 350: 1318, 375: 1319, 400: 1320, 425: 1322, 450: 1323, 475: 1324, 500: 1326, 525: 1327, 550: 1328, 575: 1329, 600: 1331, 625: 1332, 650: 1333, 675: 1334, 700: 1336, 725: 1337, 750: 1338, 775: 1340, 800: 1341},
        33: {200: 1307, 225: 1308, 250: 1309, 275: 1309, 300: 1310, 325: 1311, 350: 1312, 375: 1313, 400: 1314, 425: 1314, 450: 1315, 475: 1316, 500: 1317, 525: 1318, 550: 1319, 575: 1320, 600: 1320, 625: 1321, 650: 1322, 675: 1323, 700: 1324, 725: 1325, 750: 1326, 775: 1326, 800: 1327},
        34: {200: 1303, 225: 1304, 250: 1304, 275: 1305, 300: 1305, 325: 1306, 350: 1306, 375: 1306, 400: 1307, 425: 1307, 450: 1308, 475: 1308, 500: 1309, 525: 1309, 550: 1309, 575: 1310, 600: 1310, 625: 1311, 650: 1311, 675: 1311, 700: 1312, 725: 1312, 750: 1313, 775: 1313, 800: 1314},
        35: {200: 1300, 225: 1300, 250: 1300, 275: 1300, 300: 1300, 325: 1300, 350: 1300, 375: 1300, 400: 1300, 425: 1300, 450: 1300, 475: 1300, 500: 1300, 525: 1300, 550: 1300, 575: 1300, 600: 1300, 625: 1300, 650: 1300, 675: 1300, 700: 1300, 725: 1300, 750: 1300, 775: 1300, 800: 1300},
        36: {200: 1297, 225: 1296, 250: 1296, 275: 1295, 300: 1295, 325: 1294, 350: 1294, 375: 1294, 400: 1293, 425: 1293, 450: 1292, 475: 1292, 500: 1292, 525: 1291, 550: 1291, 575: 1290, 600: 1290, 625: 1289, 650: 1289, 675: 1289, 700: 1288, 725: 1288, 750: 1287, 775: 1287, 800: 1286},
        37: {200: 1293, 225: 1292, 250: 1292, 275: 1291, 300: 1290, 325: 1289, 350: 1288, 375: 1287, 400: 1286, 425: 1286, 450: 1285, 475: 1284, 500: 1283, 525: 1282, 550: 1281, 575: 1280, 600: 1280, 625: 1279, 650: 1278, 675: 1277, 700: 1276, 725: 1275, 750: 1275, 775: 1274, 800: 1273},
        38: {200: 1290, 225: 1289, 250: 1287, 275: 1286, 300: 1285, 325: 1283, 350: 1282, 375: 1281, 400: 1280, 425: 1278, 450: 1277, 475: 1276, 500: 1275, 525: 1273, 550: 1272, 575: 1271, 600: 1269, 625: 1268, 650: 1267, 675: 1266, 700: 1264, 725: 1263, 750: 1262, 775: 1260, 800: 1259},
        39: {200: 1286, 225: 1285, 250: 1283, 275: 1281, 300: 1280, 325: 1278, 350: 1276, 375: 1275, 400: 1273, 425: 1271, 450: 1269, 475: 1268, 500: 1266, 525: 1264, 550: 1263, 575: 1261, 600: 1259, 625: 1258, 650: 1256, 675: 1254, 700: 1252, 725: 1251, 750: 1249, 775: 1247, 800: 1246},
        40: {200: 1283, 225: 1281, 250: 1279, 275: 1277, 300: 1275, 325: 1272, 350: 1270, 375: 1268, 400: 1266, 425: 1264, 450: 1262, 475: 1260, 500: 1258, 525: 1255, 550: 1253, 575: 1251, 600: 1249, 625: 1247, 650: 1245, 675: 1243, 700: 1241, 725: 1238, 750: 1236, 775: 1234, 800: 1232},
        41: {200: 1280, 225: 1277, 250: 1275, 275: 1272, 300: 1269, 325: 1267, 350: 1264, 375: 1262, 400: 1259, 425: 1257, 450: 1254, 475: 1252, 500: 1249, 525: 1246, 550: 1244, 575: 1241, 600: 1239, 625: 1236, 650: 1234, 675: 1231, 700: 1229, 725: 1226, 750: 1224, 775: 1221, 800: 1218},
        42: {200: 1276, 225: 1273, 250: 1270, 275: 1267, 300: 1264, 325: 1261, 350: 1258, 375: 1255, 400: 1252, 425: 1249, 450: 1246, 475: 1243, 500: 1241, 525: 1238, 550: 1235, 575: 1232, 600: 1229, 625: 1226, 650: 1223, 675: 1220, 700: 1217, 725: 1214, 750: 1211, 775: 1208, 800: 1205},
        43: {200: 1273, 225: 1269, 250: 1266, 275: 1263, 300: 1259, 325: 1256, 350: 1252, 375: 1249, 400: 1246, 425: 1242, 450: 1239, 475: 1235, 500: 1232, 525: 1229, 550: 1225, 575: 1222, 600: 1218, 625: 1215, 650: 1212, 675: 1208, 700: 1205, 725: 1201, 750: 1198, 775: 1195, 800: 1191},
        44: {200: 1269, 225: 1266, 250: 1262, 275: 1258, 300: 1254, 325: 1250, 350: 1246, 375: 1243, 400: 1239, 425: 1235, 450: 1231, 475: 1227, 500: 1224, 525: 1220, 550: 1216, 575: 1212, 600: 1208, 625: 1204, 650: 1201, 675: 1197, 700: 1193, 725: 1189, 750: 1185, 775: 1181, 800: 1178},
        45: {200: 1266, 225: 1262, 250: 1258, 275: 1253, 300: 1249, 325: 1245, 350: 1241, 375: 1236, 400: 1232, 425: 1228, 450: 1224, 475: 1219, 500: 1215, 525: 1211, 550: 1207, 575: 1202, 600: 1198, 625: 1194, 650: 1190, 675: 1185, 700: 1181, 725: 1177, 750: 1173, 775: 1168, 800: 1164},
        50: {200: 1249, 225: 1243, 250: 1236, 275: 1230, 300: 1224, 325: 1217, 350: 1211, 375: 1204, 400: 1198, 425: 1192, 450: 1185, 475: 1179, 500: 1173, 525: 1166, 550: 1160, 575: 1153, 600: 1147, 625: 1141, 650: 1134, 675: 1128, 700: 1122, 725: 1115, 750: 1109, 775: 1102, 800: 1096},
    }

    # Persist the user-provided context for the running session/thread
    with detection_lock:
        detection_data['temperature_c'] = CURRENT_TEMPERATURE_C
        detection_data['htl_value'] = htl_value


    def _collect_label_centers(result, label_names):
        names = result.names if hasattr(result, "names") else result.model.names
        if result.boxes is None or len(result.boxes) == 0:
            return [], []
        boxes = result.boxes
        cls_indices = boxes.cls.cpu().numpy().astype(int)
        xyxy = boxes.xyxy.cpu().numpy()
        conf = boxes.conf.cpu().numpy() if hasattr(boxes, "conf") else None
        target_names = {label.lower() for label in label_names}
        points = []
        confidences = []
        for i, c in enumerate(cls_indices):
            if conf is not None and conf[i] < CONF_THRES:
                continue
            label = names.get(int(c), str(c)) if isinstance(names, dict) else names[int(c)]
            if str(label).lower() in target_names:
                x1, y1, x2, y2 = xyxy[i]
                cx = (x1 + x2) / 2.0
                cy = (y1 + y2) / 2.0
                points.append((cx, cy))
                confidences.append(float(conf[i]) if conf is not None else 1.0)
        return points, confidences

    def _select_reference_point(result):
        ref_points, ref_confidences = _collect_label_centers(result, REFERENCE_LABELS)
        if not ref_points:
            return None
        best_idx = max(range(len(ref_points)), key=lambda idx: ref_confidences[idx] if idx < len(ref_confidences) else 0.0)
        return ref_points[best_idx]

    def _order_points(points, confidences, reference_point):
        if not points:
            return points, confidences
        if reference_point is None:
            order = sorted(range(len(points)), key=lambda idx: points[idx][0])
        else:
            order = sorted(
                range(len(points)),
                key=lambda idx: math.hypot(points[idx][0] - reference_point[0], points[idx][1] - reference_point[1])
            )
        ordered_points = [points[idx] for idx in order]
        ordered_confidences = [confidences[idx] for idx in order]
        return ordered_points, ordered_confidences

    def get_pulley_centers(result):
        reference_point = _select_reference_point(result)
        points, confidences = _collect_label_centers(result, ("pulley",))
        points, confidences = _order_points(points, confidences, reference_point)
        return points, confidences, reference_point


    def pixel_distance(p1, p2):
        return math.hypot(p1[0] - p2[0], p1[1] - p2[1])


    def _chart_lookup_2d(temp_c, htl_val):
        """
        Look up expected X value from chart based on temperature and HTL (L/2).
        Uses bilinear interpolation for values not exactly in the chart.
        """
        if not CHART_LOOKUP:
            return None
        
        # Round temperature and HTL to nearest chart values for exact lookup
        temp_rounded = round(temp_c)
        htl_rounded = round(htl_val / 25) * 25  # Round to nearest 25
        
        # Check for exact match first
        if temp_rounded in CHART_LOOKUP and htl_rounded in CHART_LOOKUP[temp_rounded]:
            return float(CHART_LOOKUP[temp_rounded][htl_rounded])
        
        # Get available temperatures and HTL values
        temp_keys = sorted(CHART_LOOKUP.keys())
        if not temp_keys:
            return None
        
        # Find temperature bounds
        temp_lower = None
        temp_upper = None
        for t in temp_keys:
            if t <= temp_c:
                temp_lower = t
            elif t > temp_c:
                temp_upper = t
                break
        
        # Handle edge cases for temperature
        if temp_lower is None:
            temp_lower = temp_keys[0]
        if temp_upper is None:
            temp_upper = temp_keys[-1]
        
        # Get HTL values for both temperatures
        htl_keys_lower = sorted(CHART_LOOKUP[temp_lower].keys()) if temp_lower in CHART_LOOKUP else []
        htl_keys_upper = sorted(CHART_LOOKUP[temp_upper].keys()) if temp_upper in CHART_LOOKUP else []
        
        if not htl_keys_lower or not htl_keys_upper:
            return None
        
        # Find HTL bounds
        htl_lower = None
        htl_upper = None
        for h in htl_keys_lower:
            if h <= htl_val:
                htl_lower = h
            elif h > htl_val:
                htl_upper = h
                break
        
        # Handle edge cases for HTL
        if htl_lower is None:
            htl_lower = htl_keys_lower[0]
        if htl_upper is None:
            htl_upper = htl_keys_lower[-1]
        
        # Get the four corner values for bilinear interpolation
        if temp_lower == temp_upper:
            # Only interpolate along HTL axis
            if htl_lower == htl_upper:
                return float(CHART_LOOKUP[temp_lower][htl_lower])
            val_lower = CHART_LOOKUP[temp_lower][htl_lower]
            val_upper = CHART_LOOKUP[temp_lower][htl_upper]
            ratio = (htl_val - htl_lower) / (htl_upper - htl_lower) if htl_upper != htl_lower else 0
            return float(val_lower + ratio * (val_upper - val_lower))
        
        if htl_lower == htl_upper:
            # Only interpolate along temperature axis
            val_lower = CHART_LOOKUP[temp_lower][htl_lower]
            val_upper = CHART_LOOKUP[temp_upper][htl_lower]
            ratio = (temp_c - temp_lower) / (temp_upper - temp_lower) if temp_upper != temp_lower else 0
            return float(val_lower + ratio * (val_upper - val_lower))
        
        # Bilinear interpolation
        # Get four corner values
        v11 = CHART_LOOKUP[temp_lower][htl_lower]  # lower temp, lower HTL
        v12 = CHART_LOOKUP[temp_lower][htl_upper]  # lower temp, upper HTL
        v21 = CHART_LOOKUP[temp_upper][htl_lower]  # upper temp, lower HTL
        v22 = CHART_LOOKUP[temp_upper][htl_upper]  # upper temp, upper HTL
        
        # Interpolate along HTL axis first
        temp_ratio = (temp_c - temp_lower) / (temp_upper - temp_lower) if temp_upper != temp_lower else 0
        htl_ratio = (htl_val - htl_lower) / (htl_upper - htl_lower) if htl_upper != htl_lower else 0
        
        # Bilinear interpolation
        v1 = v11 + htl_ratio * (v12 - v11)  # Interpolate at lower temp
        v2 = v21 + htl_ratio * (v22 - v21)  # Interpolate at upper temp
        result = v1 + temp_ratio * (v2 - v1)  # Interpolate between temps
        
        return float(result)

    def expected_total_distance_for_temperature(temp_c, htl_val=None):
        """
        Get expected total distance based on temperature and HTL from chart.
        Falls back to simple temperature-based calculation if HTL not provided.
        """
        if htl_val is not None:
            chart_value = _chart_lookup_2d(temp_c, htl_val)
            if chart_value is not None:
                return chart_value
        
        # Fallback to simple temperature-based calculation
        delta_t = temp_c - BASE_TEMPERATURE_C
        return STANDARD_TOTAL_DISTANCE_MM - TEMPERATURE_SENSITIVITY_MM_PER_C * delta_t


    def split_total_distance(distance_mm):
        return (
            distance_mm * DISTANCE_RATIO_12,
            distance_mm * DISTANCE_RATIO_23,
        )

    def annotate_frame(frame, result):
        annotated = frame.copy()
        points, confidences, reference_point = get_pulley_centers(result)
        boxes = result.boxes
        if boxes is not None:
            cls_indices = boxes.cls.cpu().numpy().astype(int)
            xyxy = boxes.xyxy.cpu().numpy()
            conf = boxes.conf.cpu().numpy() if hasattr(boxes, "conf") else None
            names = result.names if hasattr(result, "names") else result.model.names
            for i, c in enumerate(cls_indices):
                if conf is not None and conf[i] < CONF_THRES:
                    continue
                label = names.get(int(c), str(c)) if isinstance(names, dict) else names[int(c)]
                if str(label).lower() != "pulley":
                    continue
                x1, y1, x2, y2 = map(int, xyxy[i])
                cv2.rectangle(annotated, (x1, y1), (x2, y2), (0, 255, 0), 2)
                conf_text = f"{conf[i]:.2f}" if conf is not None else "1.00"
                cv2.putText(annotated, conf_text, (x1, y1 - 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2, cv2.LINE_AA)
        if reference_point is not None:
            ref_x, ref_y = int(reference_point[0]), int(reference_point[1])
            cv2.circle(annotated, (ref_x, ref_y), 10, (0, 165, 255), 2)
            cv2.putText(annotated, "Poll", (ref_x + 10, ref_y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 165, 255), 2, cv2.LINE_AA)
        for idx, (cx, cy) in enumerate(points):
            cv2.circle(annotated, (int(cx), int(cy)), 8, (0, 255, 0), -1)
            cv2.circle(annotated, (int(cx), int(cy)), 12, (0, 255, 0), 2)
            label_text = f"P{idx + 1}"
            if idx < len(confidences):
                label_text += f" ({confidences[idx]:.2f})"
            cv2.putText(annotated, label_text, (int(cx) + 15, int(cy)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2, cv2.LINE_AA)

        dist12 = None
        dist23 = None
        total = None
        segment_distances = []
        text_y = 30

        color_palette = [
            (0, 0, 255),
            (255, 0, 0),
            (0, 255, 255),
            (255, 0, 255),
            (0, 255, 0),
        ]

        if len(points) >= 2:
            for idx in range(len(points) - 1):
                p_start = points[idx]
                p_end = points[idx + 1]
                segment_distance = pixel_distance(p_start, p_end) * MM_PER_PIXEL
                segment_distance=segment_distance*2.7 - 5 #3x right-site
                segment_distances.append(segment_distance)
                print("-----------segment_distance-----------:", segment_distance)
                color = color_palette[idx % len(color_palette)]
                cv2.line(annotated, (int(p_start[0]), int(p_start[1])), (int(p_end[0]), int(p_end[1])), color, 2)
                midpoint = (int((p_start[0] + p_end[0]) / 2), int((p_start[1] + p_end[1]) / 2))
                cv2.putText(annotated, f"{segment_distance:.2f} mm", midpoint,
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2, cv2.LINE_AA)
                cv2.putText(annotated, f"P{idx + 1}->P{idx + 2}: {segment_distance:.2f} mm", (10, text_y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2, cv2.LINE_AA)
                text_y += 28

            total = sum(segment_distances)
            # print("-----------------total---------",total)
            if total is not None:
                cv2.putText(annotated, f"P1->P{len(points)} total: {total:.2f} mm", (10, text_y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2, cv2.LINE_AA)
                text_y += 28
                if TEMPERATURE_SENSITIVITY_MM_PER_C > 0:
                    expected_total = expected_total_distance_for_temperature(CURRENT_TEMPERATURE_C, htl_value)
                    delta =  expected_total-total 
                    delta_color = (0, 255, 0) if delta >= 0 else (0, 0, 255)
                    cv2.putText(annotated, f"Δ vs standard: {delta:+.2f} mm", (10, text_y),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, delta_color, 2, cv2.LINE_AA)

            if segment_distances:
                dist12 = segment_distances[0]
                # print("-----------dist12-----------:", dist12)
            if len(segment_distances) >= 2:
                dist23 = segment_distances[1]
                # print("-----------dist23-----------:", dist23)

        height = annotated.shape[0]
        cv2.putText(annotated, f"Scale: {MM_PER_PIXEL:.3f} mm/px",
                    (10, height - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2, cv2.LINE_AA)

        return annotated, dist12, dist23, total, points, confidences, segment_distances


    def print_measurements(dist12, dist23, total):
        if dist12 is None and dist23 is None:
            print("No pulley detctions found.")
            return

        if dist12 is not None:
            print(f"Distance first->second pulley: {dist12:.3f} mm (scale {MM_PER_PIXEL} mm/px)")
        if dist23 is not None:
            print(f"Distance second->third pulley: {dist23:.3f} mm (scale {MM_PER_PIXEL} mm/px)")

        if total is not None and TEMPERATURE_SENSITIVITY_MM_PER_C > 0:
            expected_total = expected_total_distance_for_temperature(CURRENT_TEMPERATURE_C, htl_value)
            expected_dist12, expected_dist23 = split_total_distance(expected_total)
            loss_mm = expected_total - total
            print(f"Expected total distance (1->3) at {CURRENT_TEMPERATURE_C:.1f} °C (HTL {htl_value:.0f}): {expected_total:.3f} mm")
            print(f"Expected split: 1->2 = {expected_dist12:.3f} mm, 2->3 = {expected_dist23:.3f} mm")
            print(f"Measured total distance (1->3): {total:.3f} mm")
            print(f"Loss vs expected:{CURRENT_TEMPERATURE_C:.1f} °C: {expected_total:.3f} mm - {total:.3f} mm = {loss_mm:.3f} mm")
        elif dist12 is not None and dist23 is None:
            print("Only two pulleys detected; third-to-second distance not computed.")


    def main():
        global camera_cap, camera_running, detection_data, detection_lock
        
        model = _get_yolo_model()
        camera_cap = cv2.VideoCapture(0)
        camera_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        camera_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        # Optimize camera buffer size
        camera_cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        
        if not camera_cap.isOpened():
            print("Error: Could not open camera.")
            camera_running = False
            return

        camera_running = True
        print("Camera started. Detection running smoothly...")
        last_print_time = 0
        last_save_time = 0
        frame_counter = 0
        # Process every 3rd frame to reduce load (can adjust based on performance)
        PROCESS_EVERY_N_FRAMES = 3
        # Save to database every 2 seconds
        SAVE_INTERVAL_SECONDS = 2.0

        while camera_running:
            ret, frame = camera_cap.read()
            if not ret:
                print("Error: Failed to read frame from camera.")
                break
            
            frame_counter += 1
            # Skip frames to reduce processing load
            should_process = (frame_counter == 1) or (frame_counter % PROCESS_EVERY_N_FRAMES == 0)
            
            if should_process:
                # Run inference without saving frames to disk for faster startup
                results = model.predict(
                    source=frame,
                    save=False,
                    verbose=False,
                    conf=CONF_THRES,
                    iou=IOU_THRES,
                    imgsz=IMAGE_SZ,
                    agnostic_nms=False
                )
            else:
                # Use previous results or skip annotation
                results = None
                with detection_lock:
                    detection_data['latest_frame'] = frame.copy()
                    detection_data['frame_available'] = True
                    detection_data['last_update'] = time.time()

            if should_process and results:
                annotated_frame, dist12, dist23, total, points, confidences, segments = annotate_frame(frame, results[0])
                
                # Store results in shared dictionary
                with detection_lock:
                    detection_data['capture_complete'] = False
                    detection_data['dist12'] = dist12
                    detection_data['dist23'] = dist23
                    detection_data['total'] = total
                    detection_data['points'] = points
                    detection_data['confidences'] = confidences
                    detection_data['segments'] = segments
                    detection_data['pulley_count'] = len(points)
                    detection_data['temperature_c'] = CURRENT_TEMPERATURE_C
                    detection_data['htl_value'] = htl_value
                    detection_data['frame_available'] = True
                    detection_data['last_update'] = time.time()
                    
                    if total is not None and TEMPERATURE_SENSITIVITY_MM_PER_C > 0:
                        expected_total = expected_total_distance_for_temperature(CURRENT_TEMPERATURE_C, htl_value)
                        expected_dist12, expected_dist23 = split_total_distance(expected_total)
                        loss_mm = expected_total - total
                        detection_data['expected_total'] = expected_total
                        detection_data['expected_dist12'] = expected_dist12
                        detection_data['expected_dist23'] = expected_dist23
                        detection_data['loss_mm'] = loss_mm
                    else:
                        detection_data['expected_total'] = None
                        detection_data['expected_dist12'] = None
                        detection_data['expected_dist23'] = None
                        detection_data['loss_mm'] = None
                    
                    # Store the latest annotated frame for streaming
                    detection_data['latest_frame'] = annotated_frame

                current_time = time.time()
                with detection_lock:
                    capture_requested = detection_data.get('capture_requested', False)
                
                # Save to database only when user has requested capture and we have full detection
                has_full_detection = (
                    len(points) >= 3 and
                    dist12 is not None and
                    dist23 is not None and
                    total is not None
                )
                
                if (
                    capture_requested
                    and has_full_detection
                    and current_time - last_save_time >= SAVE_INTERVAL_SECONDS
                ):
                    try:
                        points_json = json.dumps([[float(p[0]), float(p[1])] for p in points]) if points else None
                        image_relative_path = save_detection_frame(annotated_frame)
                        DetectionRecord.objects.create(
                            user=detection_user,
                            pole_name=pole_name,
                            dist12=dist12,
                            dist23=dist23,
                            total=total,
                            expected_total=detection_data.get('expected_total'),
                            expected_dist12=detection_data.get('expected_dist12'),
                            expected_dist23=detection_data.get('expected_dist23'),
                            loss_mm=detection_data.get('loss_mm'),
                            pulley_count=len(points),
                            temperature_c=CURRENT_TEMPERATURE_C,
                            htl_value=htl_value,
                            points_json=points_json,
                            image_path=image_relative_path,
                        )
                        last_save_time = current_time
                        with detection_lock:
                            detection_data['capture_complete'] = True
                            detection_data['capture_image_path'] = image_relative_path
                            detection_data['capture_requested'] = False
                        print(f"Data saved to database at {timezone.now()} - full 3-pulley detection captured. Stopping camera...")
                        camera_running = False
                        break
                    except Exception as e:
                        print(f"Error saving to database: {e}")
                
                if current_time - last_print_time > 1.0 and (dist12 is not None or dist23 is not None):
                    print("-" * 60)
                    print_measurements(dist12, dist23, total)
                    last_print_time = current_time
            else:
                # Use previous annotated frame if available for smoother streaming
                with detection_lock:
                    if detection_data.get('latest_frame') is not None:
                        annotated_frame = detection_data['latest_frame'].copy()
                    else:
                        # If no previous frame, use current frame
                        annotated_frame = frame
                        if not detection_data.get('frame_available'):
                            detection_data['capture_complete'] = False
                            detection_data['pulley_count'] = 0
                            detection_data['dist12'] = None
                            detection_data['dist23'] = None
                            detection_data['total'] = None
                            detection_data['points'] = []

        if camera_cap:
            camera_cap.release()
        camera_running = False
        print("Camera stopped.")


    detection_thread = getattr(yolo_camera, "_detection_thread", None)
    if detection_thread is None or not detection_thread.is_alive():
        detection_thread = threading.Thread(target=main, daemon=True)
        detection_thread.start()
        yolo_camera._detection_thread = detection_thread

    return render(request, 'pulley_app/camera.html')

def video_stream(request):
    """Stream video frames as MJPEG"""
    def generate_frames():
        global detection_data, detection_lock, camera_running
        
        while True:
            with detection_lock:
                capture_complete = detection_data.get('capture_complete', False)
                frame_available = detection_data.get('frame_available', False)
                latest_frame = detection_data.get('latest_frame')
                camera_active = camera_running
            
            if camera_active and frame_available and latest_frame is not None:
                try:
                    frame = latest_frame.copy()
                except:
                    frame = np.zeros((480, 640, 3), dtype=np.uint8)
            else:
                # Return black frame with status text if no frame available
                frame = np.zeros((480, 640, 3), dtype=np.uint8)
                status_text = "Waiting for camera..." if not capture_complete else "Capture complete. Camera stopped."
                cv2.putText(frame, status_text, (80, 240), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
            
            # Encode frame as JPEG with optimized quality for smoother streaming
            ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
            if ret:
                frame_bytes = buffer.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
            # Slightly slower frame rate for smoother operation
            time.sleep(0.04)  # ~25 FPS for better stability
    
    return StreamingHttpResponse(generate_frames(), content_type='multipart/x-mixed-replace; boundary=frame')

def detection_results(request):
    """API endpoint to get detection results"""
    global detection_data, detection_lock
    
    with detection_lock:
        # Convert numpy arrays/tuples to lists for JSON serialization
        points = detection_data.get('points', [])
        confidences = detection_data.get('confidences', [])
        segments = detection_data.get('segments', [])
        
        # Convert points to list of lists (handles tuples, numpy arrays, etc.)
        points_list = []
        if points:
            for p in points:
                try:
                    points_list.append([float(p[0]), float(p[1])])
                except (TypeError, IndexError):
                    pass
        
        confidences_list = [float(c) for c in confidences] if confidences else []
        segments_list = [float(s) for s in segments] if segments else []
        
        data = {
            'dist12': detection_data.get('dist12'),
            'dist23': detection_data.get('dist23'),
            'total': detection_data.get('total'),
            'pulley_count': detection_data.get('pulley_count', 0),
            'temperature_c': detection_data.get('temperature_c'),
            'htl_value': detection_data.get('htl_value'),
            'expected_total': detection_data.get('expected_total'),
            'expected_dist12': detection_data.get('expected_dist12'),
            'expected_dist23': detection_data.get('expected_dist23'),
            'loss_mm': detection_data.get('loss_mm'),
            'points': points_list,
            'confidences': confidences_list,
            'segments': segments_list,
            'last_update': detection_data.get('last_update'),
            'camera_running': camera_running,
            'capture_complete': detection_data.get('capture_complete', False),
            'capture_image_path': detection_data.get('capture_image_path'),
            'capture_requested': detection_data.get('capture_requested', False),
        }
    
    return JsonResponse(data)

def stop_camera(request):
    """API endpoint to stop the camera"""
    global camera_running, camera_cap, detection_lock
    
    with detection_lock:
        camera_running = False
        if camera_cap is not None:
            camera_cap.release()
            camera_cap = None
        detection_data['capture_complete'] = False
        detection_data['capture_requested'] = False
    
    # Get final detection results
    with detection_lock:
        # Convert numpy arrays/tuples to lists for JSON serialization
        points = detection_data.get('points', [])
        confidences = detection_data.get('confidences', [])
        segments = detection_data.get('segments', [])
        
        # Convert points to list of lists (handles tuples, numpy arrays, etc.)
        points_list = []
        if points:
            for p in points:
                try:
                    points_list.append([float(p[0]), float(p[1])])
                except (TypeError, IndexError):
                    pass
        
        confidences_list = [float(c) for c in confidences] if confidences else []
        segments_list = [float(s) for s in segments] if segments else []
        
        data = {
            'dist12': detection_data.get('dist12'),
            'dist23': detection_data.get('dist23'),
            'total': detection_data.get('total'),
            'pulley_count': detection_data.get('pulley_count', 0),
            'temperature_c': detection_data.get('temperature_c'),
            'htl_value': detection_data.get('htl_value'),
            'expected_total': detection_data.get('expected_total'),
            'expected_dist12': detection_data.get('expected_dist12'),
            'expected_dist23': detection_data.get('expected_dist23'),
            'loss_mm': detection_data.get('loss_mm'),
            'points': points_list,
            'confidences': confidences_list,
            'segments': segments_list,
            'last_update': detection_data.get('last_update'),
            'camera_running': False,
            'capture_complete': detection_data.get('capture_complete', False),
            'capture_image_path': detection_data.get('capture_image_path'),
            'capture_requested': detection_data.get('capture_requested', False),
            'success': True
        }
    
    return JsonResponse(data)
# Config (mirrors predict.py / video.py)


def request_capture(request):
    """Flag that the next full detection should be saved"""
    global detection_data, detection_lock
    with detection_lock:
        detection_data['capture_requested'] = True
        detection_data['capture_complete'] = False
    return JsonResponse({'capture_requested': True})








#  Download recorded detections as CSV,pdf, or excel
# import csv
# from django.http import HttpResponse
# from  pulley_app.models import PulleyDetection,DetectionRecord, CustomUser
# from openpyxl import Workbook
# from reportlab.pdfgen import canvas

# def export_csv(request):
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="PulleyDetection.csv"'

#     writer = csv.writer(response)
#     writer.writerow(['ID','user', 'pole_name', 'uploaded_image', 'result_image','temperature_c', 'htl_value','dist_p1_p2', 'dist_p2_p3', 'total_distance', 'expected_total', 'loss_mm', 'distances', 'created_at'])

#     for detect in PulleyDetection.objects.all():
#         writer.writerow([detect.id,detect.user, detect.pole_name, detect.uploaded_image,detect.result_image, detect.temperature_c, detect.htl_value, detect.dist_p1_p2, detect.dist_p2_p3, detect.total_distance, detect.expected_total, detect.loss_mm, detect.distances, detect.created_at])

#     return response
import csv
from django.http import HttpResponse
from pulley_app.models import PulleyDetection

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="PulleyDetection.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'User', 'Pole Name', 'Uploaded Image', 'Result Image',
        'Temperature (°C)', 'HTL Value', 'Dist P1-P2', 'Dist P2-P3',
        'Total Distance', 'Expected Total', 'Loss (mm)',
        'Distances', 'Created At'
    ])

    detections = PulleyDetection.objects.filter(
        user=request.user
    ).order_by('-id')

    for detect in detections:
        writer.writerow([
            detect.id,
            detect.user.username if detect.user else '',
            detect.pole_name,
            detect.uploaded_image.url if detect.uploaded_image else '',
            detect.result_image.url if detect.result_image else '',
            detect.temperature_c,
            detect.htl_value,
            detect.dist_p1_p2,
            detect.dist_p2_p3,
            detect.total_distance,
            detect.expected_total,
            detect.loss_mm,
            detect.distances,
            detect.created_at
        ])

    return response



# def export_csv1(request):
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="DetectionRecord.csv"'

#     writer = csv.writer(response)
#     writer.writerow(['ID','user', 'pole_name', 'dist12', 'dist23', 'total', 'expected_total', 'expected_dist12', 'expected_dist23', 'loss_mm', 'pulley_count', 'temperature_c', 'htl_value', 'points_json', 'image_path', 'timestamp'])

#     for detect in DetectionRecord.objects.all():
#         writer.writerow([detect.id,detect.user, detect.pole_name, detect.dist12, detect.dist23, detect.total, detect.expected_total, detect.expected_dist12, detect.expected_dist23, detect.loss_mm, detect.pulley_count, detect.temperature_c, detect.htl_value, detect.points_json, detect.image_path, detect.timestamp])

#     return response

import csv
from django.http import HttpResponse
from .models import DetectionRecord,PulleyDetection

def download_record_csv(request, record_id):
    record = get_object_or_404(PulleyDetection, id=record_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="data_record_{record.id}.csv"'
    )

    writer = csv.writer(response)

    # CSV Header
    writer.writerow([
        'ID',
        'User',
        'Pole Name',
        'Upload Image',
        'Result Image',
        'Temperature (°C)',
        'HTL Values',
        'Dist-p1-p2',
        'Dist-p2-p3',
        'Total Dict',
        'Expected_Total',
        'Loss_mm',
        'Distance',
        'Created At'
    ])

    # CSV Row
    writer.writerow([
        record.id,
        record.user.username if record.user else 'N/A',
        record.pole_name,
        record.uploaded_image,
        record.result_image,
        record.temperature_c,
        record.htl_value,
        record.dist_p1_p2,
        record.dist_p2_p3,
        record.total_distance,
        record.expected_total,
        record.loss_mm,
        record.distances,
        record.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    ])

    return response



def download_record_csv2(request, record_id):
    record = get_object_or_404(DetectionRecord, id=record_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="data_record_{record.id}.csv"'
    )

    writer = csv.writer(response)

    # CSV Header
    writer.writerow([
        'ID',
        'User',
        'Pole Name',
        'Timestamp',
        'Dist12',
        'Dist23',
        'Total Dict',
        'Expected_total',
        'Expected_dist12',
        'Expected_dist23',
        'Loss_mm',
        'pulley_count',
        'Temperature_c',
        'HTL Values',
        'image_path',
    ])

    # CSV Row
    writer.writerow([
        record.id,
        record.user.username if record.user else 'N/A',
        record.pole_name,
        record.timestamp,
        record.dist12,
        record.dist23,
        record.total,
        record.expected_total,
        record.expected_dist12,
        record.expected_dist23,
        record.expected_total,
        record.loss_mm,
        record.pulley_count,
        record.temperature_c,
        record.htl_value,
       
        
    ])

    return response







def export_csv1(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="DetectionRecord.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'User', 'Pole Name',
        'Dist12', 'Dist23', 'Total',
        'Expected Total', 'Expected Dist12', 'Expected Dist23',
        'Loss mm', 'Pulley Count',
        'Temperature C', 'HTL Value',
        'Points JSON', 'Image Path', 'Timestamp'
    ])

    records = DetectionRecord.objects.filter(
        user=request.user
    ).order_by('-id')

    for detect in records:
        writer.writerow([
            detect.id,
            detect.user.username if detect.user else '',
            detect.pole_name,
            detect.dist12,
            detect.dist23,
            detect.total,
            detect.expected_total,
            detect.expected_dist12,
            detect.expected_dist23,
            detect.loss_mm,
            detect.pulley_count,
            detect.temperature_c,
            detect.htl_value,
            detect.points_json,
            detect.image_path,
            detect.timestamp
        ])

    return response


from openpyxl import Workbook # type: ignore
from django.http import HttpResponse
from .models import PulleyDetection


def export_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PulleyDetection"

    # HEADER
    sheet.append([
        'ID', 'User', 'Pole Name', 'Uploaded Image', 'Result Image',
        'Temperature C', 'HTL Value', 'Dist P1 P2', 'Dist P2 P3',
        'Total Distance', 'Expected Total', 'Loss mm',
        'Distances', 'Created At'
    ])

    detections = PulleyDetection.objects.filter(
        user=request.user
    ).order_by('-id')

    for d in detections:
        sheet.append([
            d.id,
            d.user.username if d.user else '',
            d.pole_name or '',
            d.uploaded_image.url if d.uploaded_image else '',
            d.result_image.url if d.result_image else '',
            d.temperature_c,
            d.htl_value,
            d.dist_p1_p2,
            d.dist_p2_p3,
            d.total_distance,
            d.expected_total,
            d.loss_mm,
            str(d.distances),
            d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="PulleyDetection.xlsx"'

    workbook.save(response)
    return response


# def export_excel(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "PulleyDetection"

#     sheet.append(['ID', 'user',' uploaded_image', 'result_image','temperature_c', 'htl_value','dist_p1_p2', 'dist_p2_p3', 'total_distance', 'expected_total', 'loss_mm', 'distances', 'created_at'])

#     for detect in PulleyDetection.objects.all():
#         sheet.append([detect.id, detect.user, detect.uploaded_image,detect.result_image, detect.temperature_c, detect.htl_value, detect.dist_p1_p2, detect.dist_p2_p3, detect.total_distance, detect.expected_total, detect.loss_mm, detect.distances, detect.created_at])

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="employee.xlsx"'

#     workbook.save(response)
#     return response



# from openpyxl import Workbook
# from django.http import HttpResponse
# from .models import PulleyDetection


# def export_excel(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "PulleyDetection"

#     # HEADER
#     sheet.append([
#         'ID', 'User', 'Pole Name', 'Uploaded Image', 'Result Image',
#         'Temperature C', 'Htl Value', 'Dist P1 P2', 'Dist P2 P3',
#         'Total Distance', 'Expected Total', 'Loss mm',
#         'Distances', 'Created At'
#     ])

#     detections = PulleyDetection.objects.all()

#     for d in detections:
#         sheet.append([
#             d.id,
#             str(d.user),                     # FIXED
#             d.pole_name or '',               # FIXED
#             str(d.uploaded_image),           # FIXED
#             str(d.result_image),             # FIXED
#             d.temperature_c,
#             d.htl_value,
#             d.dist_p1_p2,
#             d.dist_p2_p3,
#             d.total_distance,
#             d.expected_total,
#             d.loss_mm,
#             str(d.distances),                # FIXED
#             d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else ""
#         ])

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="PulleyDetection.xlsx"'

#     workbook.save(response)
#     return response

# def export_excel1(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "DetectionRecord"

#     # HEADER
#     sheet.append([
#         # 'id','user', 'pole_name', 'timestamp', 'dist12', 'dist23', 'total', 'expected_total', 'expected_dist12', 'expected_dist23', 'loss_mm', 'pulley_count', 'temperature_c', 'htl_value', 'points_json', 'image_path', 'ordering', ('ID', 'User', 'Pole Name', 'dist12', 'dist23', 'total', 'expected_total', 'expected_dist12', 'expected_dist23', 'loss_mm', 'pulley_count', 'temperature_c', 'htl_value', 'points_json', 'image_path', 'timestamp'
#         'ID', 'User', 'Pole Name', 'dist12', 'dist23', 'total', 
#         'expected_total', 'expected_dist12', 'expected_dist23', 
#         'loss_mm', 'pulley_count', 'temperature_c', 'htl_value',
#         'timestamp'
#     ])

#     detections = DetectionRecord.objects.all()

#     for d in detections:
#         sheet.append([
#             d.id,
#             str(d.user),                     # FIXED
#             d.pole_name or '',               # FIXED
#             # str(d.uploaded_image),           # FIXED
#             # str(d.result_image),             # FIXED
#             d.temperature_c,
#             d.htl_value,
#             d.dist12,
#             d.dist23,
#             d.total,
#             d.expected_total,
#             d.loss_mm,
#             d.pulley_count,
#             d.temperature_c,
#             d.htl_value,
#             # d.points_json,
#             # d.image_path,
#             d.timestamp.strftime("%Y-%m-%d %H:%M:%S") if d.timestamp else "",
#             # d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else ""
#         ])

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="PulleyDetection.xlsx"'

#     workbook.save(response)
#     return response
from openpyxl import Workbook # type: ignore
from django.http import HttpResponse
from .models import DetectionRecord


def export_excel1(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DetectionRecord"

    # HEADER (matches data exactly)
    sheet.append([
        'ID', 'User', 'Pole Name',
        'Dist12', 'Dist23', 'Total',
        'Expected Total', 'Expected Dist12', 'Expected Dist23',
        'Loss mm', 'Pulley Count',
        'Temperature C', 'HTL Value',
        'Timestamp'
    ])

    detections = DetectionRecord.objects.filter(
        user=request.user
    ).order_by('-id')

    for d in detections:
        sheet.append([
            d.id,
            d.user.username if d.user else '',
            d.pole_name or '',
            d.dist12,
            d.dist23,
            d.total,
            d.expected_total,
            d.expected_dist12,
            d.expected_dist23,
            d.loss_mm,
            d.pulley_count,
            d.temperature_c,
            d.htl_value,
            d.timestamp.strftime("%Y-%m-%d %H:%M:%S") if d.timestamp else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="DetectionRecord.xlsx"'

    workbook.save(response)
    return response



from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph # type: ignore
from reportlab.lib.pagesizes import landscape, A4 # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # type: ignore
from reportlab.lib.units import mm # type: ignore
from django.http import HttpResponse
from .models import PulleyDetection


def export_pdf(request):

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="PulleyDetection_Report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=10,
        rightMargin=10,
        topMargin=10,
        bottomMargin=10
    )

    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.fontSize = 6      # Smaller text → fit more columns
    styleN.leading = 7       # Tight line spacing

    # Load DB
    data_rows = PulleyDetection.objects.all()

    # Detect all fields
    fields = [f.name for f in PulleyDetection._meta.get_fields() if f.concrete]

    # Create readable headers
    header = [f.replace("_", " ").title() for f in fields]

    data = [header]

    # Build rows
    for row in data_rows:
        row_data = []
        for field in fields:
            value = getattr(row, field)

            # Wrap long content
            value = "" if value is None else str(value)
            row_data.append(Paragraph(value, styleN))

        data.append(row_data)

    # PROFESSIONAL COLUMN WIDTH HANDLING  
    col_widths = []
    for field in fields:
        f = field.lower()

        if "image" in f:
            col_widths.append(35 * mm)  # For long filenames
        elif "distance" in f or "result" in f or "loss" in f:
            col_widths.append(40 * mm)
        elif "created" in f or "updated" in f:
            col_widths.append(30 * mm)
        else:
            col_widths.append(22 * mm)  # Default width

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Table styling
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Table font for body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),

        # Row height minimum
        ('MINROWHEIGHT', (0, 0), (-1, -1), 8),

        # Cell padding smaller (more compact)
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),

        # Border grid
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    doc.build([table])
    return response




